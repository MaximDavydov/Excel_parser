# pip install openpyxl

import openpyxl
import os


# os.mkdir("documents")

# Define variable to load the wookbook
wookbook = openpyxl.load_workbook("Задание.xlsx")

print(wookbook.get_sheet_names())
# Define variable to read the active sheet:
worksheet = wookbook.active
# Iterate the loop to read the cell values
doc_list = []
for i in range(1, worksheet.max_row):
    name_list = []

    for col in worksheet.iter_cols(1, worksheet.max_column):
        name_list.append(col[i].value)
    if name_list[1] == 'акт,счет':
        doc_list = name_list[1].split(',')
        path_str_1 = f'//documents/{name_list[0]}_{doc_list[0]}_{name_list[2]}'
        my_file_1 = open(path_str_1, "a+")

        path_str_2 = f'//documents/{name_list[0]}_{doc_list[1]}_{name_list[2]}'
        my_file_2 = open(path_str_2, "a+")
    else:
        path_str_2 = f'//documents/{name_list[0]}_{name_list[1]}_{name_list[2]}'
        my_file_2 = open(path_str_2, "a+")

doc_list = os.listdir('//documents')
# print(doc_list)




